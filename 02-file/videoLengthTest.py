import os;
import natsort;
from openpyxl import Workbook;
from moviepy.editor import VideoFileClip;



# 每个文件夹对象属性:
# 开始行数 / 子行数 / 总时长

BEGIN_ROW = "beginRow";
ROW_TOTAL = "rowTotal";
TOTAL_LENGTH = "totalLength";

path = "/Users/luoran/Documents/01-course/01-aws/main/";
suffix = ".mp4";
currentRow = 1; # 第一行为表头

fileList=os.listdir(path);
fileList = natsort.natsorted(fileList);

def operEachFile(subFile, subPath, fileInfo, outws):
    print("开始处理单个文件,入参为{}", subFile)
    if not (subFile.endswith(suffix)):
        return;
    global currentRow;
    currentRow += 1;
    # 统计行数
    fileInfo[ROW_TOTAL] += 1;
    # 名称写入excel
    outws['B{}'.format(currentRow)] = subFile; 
    # 获取长度写入excel]
    video = VideoFileClip(subPath + '/' + subFile);
    duration = video.duration;
    outws['C{}'.format(currentRow)] = duration;
    fileInfo[TOTAL_LENGTH] += duration;

def main():
    outwb = Workbook();
    outws = outwb.worksheets[0];
    # 输出表头
    outws.append(['文件夹','视频','长度','总长'])

    for file in fileList:
        if (file.startswith(".DS_Store")):
            continue;
        
        fileInfo = {
            BEGIN_ROW:currentRow + 1,
            ROW_TOTAL:0,
            TOTAL_LENGTH:0
        }
    
        subPath = path + file;
        subFileList = os.listdir(subPath);
        subFileList = natsort.natsorted(subFileList);
        # 一级文件夹名写入excel 合并行数 = 子文件数

        # 遍历处理子文件
        for subFile in subFileList:
            operEachFile(subFile, subPath, fileInfo, outws);

        # 合并单元格写入 文件夹名 与 总时长
        if (fileInfo[ROW_TOTAL]<1):
            outws['A{}'.format(fileInfo[BEGIN_ROW])] = file;
            outws['D{}'.format(fileInfo[BEGIN_ROW])] = 0;
        else:
            outws.merge_cells('A{}:A{}'.format(fileInfo[BEGIN_ROW], fileInfo[BEGIN_ROW]+fileInfo[ROW_TOTAL]-1));
            outws['A{}'.format(fileInfo[BEGIN_ROW])] = file;
            outws.merge_cells('D{}:D{}'.format(fileInfo[BEGIN_ROW], fileInfo[BEGIN_ROW]+fileInfo[ROW_TOTAL]-1));
            outws['D{}'.format(fileInfo[BEGIN_ROW])] = fileInfo[TOTAL_LENGTH];            

    # 保存为文件
    outwb.save('result.xlsx');
    print("save successfully")

if __name__ == "__main__":
    main()